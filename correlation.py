
import re
import openpyxl as xl 
import pandas as pd
import numpy as np 
import warnings
import tkinter as ttk 

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')


#TO-DO // comment gérer les excel ? sans les télécharger : -les host- 
#TO-DO // voir la date et supprimer les semaines déjà passer 

#"Staff_1":"10h-11h30", #8 à 14
#"Staff_2":"11h30-13h", #14 à 20
#"Staff_3":"13h-14h30", #20 à 26
#"Staff_4":"14h30-16h" # 26 à 32

planning_staff = [list(range(8,15)),list(range(14,21)),list(range(20,27)),list(range(26,33))]

path = "/Users/eliotwalter/Desktop/coursC.xlsx"


wb = xl.load_workbook(path)
del wb['EC-Planning']

df1 = pd.read_excel(path)
print(wb.sheetnames)



day_to_number = {"lundi":"1","mardi":"2","mercredi":"3","jeudi":"4","vendredi":"5"}


#récupéré le sheet name, le montrer à l'utilisateur et le laisser choisir, l'utiliser dans le programme 

def week_to_choose(week_choosen):
    
    data = week_choosen

    titres_de_feuilles = wb.sheetnames
    regex_pattern = re.compile(fr".*{re.escape(data)}.*")

    for titre in titres_de_feuilles:
        if regex_pattern.match(titre):
            #print(f"'{data}' est inclus dans le titre : {titre}")
            pass
    indice = titres_de_feuilles.index(titre)
    return indice #gérer le fait qu'il faut rentrer la date et le mois ou un choix en tkinter ou géré l'année 


def search_dispo(date_choosen): #rajouter nom et semaine 
    indice = week_to_choose(n)
    df1 = pd.read_excel(path,indice) #multithread à faire 

    print("C'est la semaine de "+df1.columns.tolist()[0] + ".")


    list_staff_horaires = list(range(7,33))

    horaires1 = df1.iloc[1:57,0].dropna() #garder ces informations en mémoire 
    horaires2 = df1.iloc[1:57,7].dropna()
    horaires_total = pd.concat([horaires1,horaires2]).sort_values(ascending=True)
    horaires_staff_total = horaires_total.iloc[list_staff_horaires]

    numero_jour = int(day_to_number[date_choosen]) #remplacer 
    day = df1.iloc[0:,[numero_jour]]
    day_index = day.dropna().index.tolist()
    print(day_index)
    try: 
        horaire_début_cours = min(day_index)
        horaire_fin_cours = max(day_index)+5
    except: 
        horaire_début_cours = 0
        horaire_fin_cours = 0
    horaires_cours = list(range(horaire_début_cours,horaire_fin_cours))

    horaires_de_staff = horaires_staff_total.index.tolist() #horaires de staff 
    horaires_dispo = list(set(horaires_de_staff) | set(horaires_cours))

    # Utilisez l'opérateur de soustraction pour obtenir les éléments qui ne sont présents que dans l'une des listes
    horaires_dispo = [x for x in horaires_dispo if (x not in horaires_cours or x not in horaires_de_staff) and  8 <= x <= 32] #horaires de staff dispo 

    for staff in planning_staff:
        if all(heures in horaires_dispo for heures in staff):
            print(f"tu peux staffer pour le {staff}") #changer ce que resort le programme à travers les heures : pas besoin 
        else: 
            print(f"ah jcrois tu peux pas staff pour ce moment : {staff}") 



#print(df1.shape)

#le planning est considéré comme étant le titre, donc les jours commencent en 0, ensuite les données en 1 
#exclure les pauses 
#faire correspondre les cours avec les heures associés 
    #si langue : cas spécial 

#for y in horaires_dispo: 
   # print(horaires_total.iloc[y])
n = input("quel semaine tu veux : ")
jour = input("quel jour veux-tu :").lower()
search_dispo(jour)
